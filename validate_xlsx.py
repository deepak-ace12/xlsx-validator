from email import header
from optparse import Values
import yaml
import sys
import copy
from collections import defaultdict
from openpyxl.reader.excel import load_workbook
from validators_pd import *

import os, psutil


ERRORS = defaultdict(list)

rv = RequiredValidator()
ov = OptionValidator()
tv = TypeValidator()
dtv = DateTimeValidator()
nnv = NonNegativeValidator()
ev = EmailValidator()
rgv = RegexValidator()
lv = LengthValidator()


v_obj = {
    "RequiredValidator": rv,
    "OptionValidator": ov,
    "TypeValidator": tv,
    "DateTimeValidator": dtv,
    "NonNegativeValidator": nnv,
    "RegexValidator": rgv,
    "EmailValidator": ev,
    "LengthValidator": lv
}

def is_valid_cell(valdn_type, cell, sheet, column_header):
    for typ, data in valdn_type.items():
        metadata = {
            "header": column_header,
            "cell": cell.coordinate,
        }
        # validating_class = eval(typ) #getattr(sys.modules[__name__], typ)
        validator = v_obj.get(typ)
        try:
            validator.validate(cell.value, data)
        except Exception as ex:
            metadata["error"] = ex.args[0]
            ERRORS[sheet].append(metadata)


def set_config(yaml_config):
    """
    Takes the config yaml file and converts it into a dictionary.
    """
    with open(yaml_config, "r") as yml:
        config = yaml.safe_load(yml)
        return config


def validate(config, worksheet):
    columns_to_validate = config.get("validations").get("columns") or []
    must_have_columns = config.get("must_have_columns")
    # ********************** COLUMN_CASES ******************* #
    column_letter_to_header = {}
    for row in worksheet.iter_rows(max_row=1):
        for cell in row:
            if cell.value:
                column_letter_to_header[cell.column_letter] = cell.value
    # ********************** COLUMN_CASES ******************* #
    if not set(must_have_columns).issubset(set(column_letter_to_header.values())):
        raise Exception(
            {
                "sheet": worksheet.title,
                "error": "Sheet {sheet} must have column(s) {missing_columns}".format(
                    sheet=worksheet.title,
                    missing_columns=", ".join(
                        list(
                            set(must_have_columns)
                            - set(column_letter_to_header.values())
                        )
                    ),
                ),
            }
        )
    start_row = 1
    if config.get("iterate_by_header_name"):
        start_row = 2
    else:
        for key, _ in column_letter_to_header.items():
            column_letter_to_header[key] = key
    for row in worksheet.iter_rows(min_row=start_row, max_col=len(column_letter_to_header.keys())):
        if not all([cell.value is None for cell in row]):
            for cell in row:
                column_header = column_letter_to_header[cell.column_letter]
                if column_header in config.get("read_as_string", []):
                    cell.value = str(cell.value)
                if column_header in config.get("read_as_int", []):
                    cell.value = int(cell.value)
                if column_header in config.get("exclude", []):
                    continue
                if column_header in columns_to_validate:
                    for valdn_type in columns_to_validate[column_header]:
                        is_valid_cell(
                            valdn_type=valdn_type,
                            cell=cell,
                            sheet=worksheet.title,
                            column_header=column_header
                        )

                elif config.get("validations").get("default"):
                    is_valid_cell(
                        valdn_type=config.get("validations").get("default")[0],
                        cell=cell,
                        sheet=worksheet.title,
                        column_header=column_header
                    )

    # for error in ERRORS:
    #     print(error)  # TODO: do something about error logging

def validate_excel(xlsx_filepath, yaml_filepath):

    try:
        import time
        t1 = time.time()
        workbook = load_workbook(xlsx_filepath)
        sheets = workbook.sheetnames
        config = set_config(yaml_filepath)
        config_sheets = config.get("sheets")
        t2 = time.time()
        print("loading", (t2-t1))
        if not set(config_sheets).issubset(set(sheets)):
            raise Exception(
            {
                "error": "The uploaded file must have sheet(s) {missing_sheets}".format(
                    missing_sheets=", ".join(
                        list(
                            set(config_sheets)
                            - set(sheets)
                        )
                    ),
                ),
            }
        )
        for sheet in sheets:
            if config.get(sheet):
                t11 = time.time()
                worksheet = workbook[sheet]
                t22 = time.time()
                print("sheet", (t22-t11))
                validate(config.get(sheet), worksheet)
                t33 = time.time()
                print("vals", (t33-t22))
    except Exception as e:
        sys.exit(e.args[0])
    
    print(ERRORS)


if __name__ == "__main__":
    import time
    t1 = time.time()
    # xlsx_filepath = "/Users/I1597/Downloads/ucc_data_columns_final.xlsx"
    # yaml_filepath = "/Users/I1597/Documents/repositories/excel_validator/thn_final_validator.yml"
    xlsx_filepath = "/Users/I1597/Documents/repositories/excel_validator/one_lakh_record.xlsx"
    yaml_filepath = "/Users/I1597/Documents/repositories/excel_validator/sales_record.yml"

    validate_excel(xlsx_filepath, yaml_filepath)
    t2 = time.time()
    print("Total Time", (t2 - t1))
print("RAM USED",(psutil.Process().memory_info().rss / 1024 ** 2) )
print("HARD DISK USED", (psutil.Process().memory_info().vms / 1024 ** 2))
print("CPU USED", psutil.cpu_percent(), "%")
