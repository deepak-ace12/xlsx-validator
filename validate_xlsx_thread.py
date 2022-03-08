import sys
import yaml
from openpyxl.reader.excel import load_workbook
from validators import (
    RequiredValidator,
    TypeValidator,
    LengthValidator,
    RegexValidator,
    EmailValidator,
    OptionValidator,
    DateTimeValidator,
    ExcelDateValidator,
    NonNegativeValidator,
    ComparatorValidator,
)
from multiprocessing.pool import ThreadPool as Pool

pool = Pool(10)
column_letter_to_header = {
    "A": "Region",
    "B": "Country",
    "C": "Item Type",
    "D": "Sales Channel",
    "E": "Order Priority",
    "F": "Order Date",
    "G": "Order ID",
    "H": "Ship Date",
    "I": "Units Sold",
    "J": "Unit Price",
    "K": "Unit Cost",
    "L": "Total Revenue",
    "M": "Total Cost",
    "N": "Total Profit",
}
errors = []
config = {
    "excludes": [],
    "iterate_by_header_name": True,
    "validations": {
        "default": [{"Required": {"error_msg": "Cell Value can not be blank."}}],
        "columns": {
            "Sales Channel": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "Option": {
                        "options": ["Online", "Offline"],
                        "error_msg": "Cell value must be a valid option.",
                        "case_sensitive": False,
                    },
                }
            ],
            "Order Priority": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "Option": {
                        "options": ["C", "H", "L", "M"],
                        "case_sensitive": False,
                        "error_msg": "Invalid option.",
                    },
                }
            ],
            "Order Date": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "Date": {
                        "format": "%m/%d/%Y",
                        "error_msg": "Date format must be mm/dd/yyyy.",
                    },
                }
            ],
            "Units Sold": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "NonNegative": {
                        "type": "int",
                        "error_msg": "Cell value must be a non negative value.",
                    },
                }
            ],
            "Unit Price": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "Type": {
                        "type": "float",
                        "error_msg": "Cell value must be a decimal value.",
                    },
                },
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "NonNegative": {
                        "type": "int",
                        "error_msg": "Cell value must be a non negative value.",
                    },
                },
            ],
            "Unit Cost": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "Type": {
                        "type": "float",
                        "error_msg": "Cell value must be a decimal value.",
                    },
                },
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "NonNegative": {
                        "type": "int",
                        "error_msg": "Cell value must be a non negative value.",
                    },
                },
            ],
            "Total Revenue": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "Type": {
                        "type": "float",
                        "error_msg": "Cell value must be a decimal value.",
                    },
                },
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "NonNegative": {
                        "type": "int",
                        "error_msg": "Cell value must be a non negative value.",
                    },
                },
            ],
            "Total Cost": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "Type": {
                        "type": "float",
                        "error_msg": "Cell value must be a decimal value.",
                    },
                },
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "NonNegative": {
                        "type": "int",
                        "error_msg": "Cell value must be a non negative value.",
                    },
                },
            ],
            "Total Profit": [
                {
                    "Required": {"error_msg": "Cell Value can not be blank."},
                    "Type": {
                        "type": "float",
                        "error_msg": "Cell value must be a decimal value.",
                    },
                }
            ],
        },
    },
}
columns_to_validate = []


def is_valid_cell(valdn_type, value, coordinates, errors):
    classmap = {
        "Required": RequiredValidator,
        "Type": TypeValidator,
        "Length": LengthValidator,
        "Regex": RegexValidator,
        "Email": EmailValidator,
        "Option": OptionValidator,
        "Date": DateTimeValidator,
        "ExcelDate": ExcelDateValidator,
        "Comparator": ComparatorValidator,
        "NonNegative": NonNegativeValidator,
        "Datetime": DateTimeValidator,
    }

    for typ, data in valdn_type.items():
        validator = classmap[typ](data)  # creating the object with the parameters
        try:
            validator.validate(value)
        except Exception as ex:
            coordinates["Error"] = ex.args[0]
            if coordinates not in errors:
                errors.append(coordinates)


def set_config(yaml_config, yaml_validator_cls):
    """
    Takes the config yaml file and converts it into a dictionary.
    """
    with open(yaml_config, "r") as yml:
        config = yaml.safe_load(yml).get(yaml_validator_cls)
        # config["default"] = config.get("validations").get("default")[0] or None
        print(config)
        return config


def execute_threads(row):
    for cell in row:
        column_header = column_letter_to_header[cell.column_letter]
        coordinates = {
            "Header": column_header,
            "Cell": cell.coordinate,
        }
        if column_header in config.get("excludes", []):
            continue
        if column_header in columns_to_validate:
            for valdn_type in columns_to_validate[column_header]:
                is_valid_cell(valdn_type, cell.value, coordinates, errors)

        elif config.get("validations").get("default"):
            is_valid_cell(
                config.get("validations").get("default")[0],
                cell.value,
                coordinates,
                errors,
            )


def validate(config, worksheet):
    columns_to_validate = config.get("validations").get("columns")

    # ********************** COLUMN_CASES ******************* #
    column_letter_to_header = {}
    for row in worksheet.iter_rows(max_row=1):
        for cell in row:
            if cell.value:
                column_letter_to_header[cell.column_letter] = cell.value
    # ********************** COLUMN_CASES ******************* #
    print(column_letter_to_header)
    start_row = 1
    if config.get("iterate_by_header_name"):
        start_row = 2
    else:
        for key, _ in column_letter_to_header.items():
            column_letter_to_header[key] = key
    p_outputs = pool.map(
        execute_threads,
        worksheet.iter_rows(min_row=start_row, max_col=worksheet.max_column),
    )

    for error in errors:
        print(error)  # TODO: do something about error logging


def validate_excel(xlsx_filepath, yaml_filepath):
    try:
        workbook = load_workbook(xlsx_filepath, read_only=True)
        sheets = workbook.sheetnames
        for sheet in sheets:
            config = set_config(yaml_filepath, sheet)
            worksheet = workbook[sheet]
            validate(config, worksheet)
    except Exception as e:
        sys.exit("Error occured: " + str(e))


if __name__ == "__main__":
    import time

    t1 = time.time()
    xlsx_filepath = (
        "/Users/I1597/Documents/repositories/excel_validator/one_lakh_record.xlsx"
    )
    yaml_filepath = (
        "/Users/I1597/Documents/repositories/excel_validator/sales_record.yml"
    )
    validate_excel(xlsx_filepath, yaml_filepath)
    t2 = time.time()
    print("Total Time", (t2 - t1))
