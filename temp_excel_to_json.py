from openpyxl.reader.excel import load_workbook
from collections import defaultdict
import json

file_path = "/Users/I1597/Downloads/ucc_data_2.xlsx"

workbook = load_workbook(file_path)
sheets = ["Patients", "Care Team", "Care Team Members"]
sheet_data = defaultdict(list)
for sheet in sheets:
    column_letter_to_header = {}
    worksheet = workbook[sheet]
    for row in worksheet.iter_rows(max_row=1):
        for cell in row:
            if cell.value:
                column_letter_to_header[cell.column_letter] = cell.value

    print(sheet)
    for row in worksheet.iter_rows(min_row=2, max_col=len(column_letter_to_header.keys())):
        temp = {}
        if not all([cell.value == None for cell in row]):
            for cell in row:
                temp[column_letter_to_header.get(cell.column_letter)] = cell.value
        sheet_data[sheet].append(temp)
    # print(sheet_data)
print(sheet_data)

patients = sheet_data.get("Patients")
care_team = sheet_data.get("Care Team")
care_team_members = sheet_data.get("Care Team Members")

sample_patient_data = {}
for patient in patients:
    phone_number = patient.get("Phone Number")
    temp = {}
    temp["patient"] = {}
    temp["patient"]["empi"] = patient.get("EMPI")
    temp["patient"]["firstName"] = patient.get("First Name")
    temp["patient"]["lastName"] = patient.get("Last name")
    temp["patient"]["address"] = patient.get("Address")
    care_team_id = patient.get("Care Team ID")
    for team in care_team:
        if care_team_id == team.get("Care Team ID"):
            temp["primaryCMA"] = {
                "name": team.get("CPP or RN"),
                "email": team.get("CPP Id"),
            }
    for member in care_team_members:
        if care_team_id == member.get("Care Team ID"):
            temp["alternateCMA"] = {
                "name": member.get("Name"),
                "email": member.get("Care Team Member Id"),
            }
    sample_patient_data[str(int(phone_number))] = temp

with open("data_sample.json", "w") as f:
    json.dump(sample_patient_data, f)
with open("data_sample.json", "w") as f:
    json.dump(sample_patient_data, f)

print()


def get_parsed_file_data(self, sheets):
    sheet_data = {}
    numeric_to_str = [
        "Phone Number",
        "CPP Phone",
    ]
    for sheet in sheets:
        for column in numeric_to_str:
            if sheet.get(column):
                sheet = sheet.apply(lambda x: str(x))
        sheet_dict = sheet.to_dict(orient="records")
        sheet_data[sheet] = sheet_dict
    return sheet_data